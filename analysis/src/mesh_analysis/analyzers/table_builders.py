# TODO: Cleanup. AI used.

import logging
from dataclasses import dataclass
from io import StringIO
from typing import Any, Dict, List, Self

import pandas as pd
import plotly.graph_objects as go
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.tree import Tree
from src.mesh_analysis.analyzers.analyzer import AnalysisResult
from tabulate import tabulate

logger = logging.getLogger(__name__)
sns.set_theme()


@dataclass
class Experiment:
    name: str
    tests: List[AnalysisResult]


def print_tree(experiments: List[Dict[str, Any]]):
    tree = Tree("Experiments")
    for exp in experiments:
        exp_node = tree.add(exp["name"])
        for test in exp["tests"]:
            test_node = exp_node.add(f"[bold]{test.name}:[/] {test.status}")
            if test.intermediates:
                for key, value in test.intermediates.items():
                    test_node.add(f"  {key}: {value}")
    console = Console(file=StringIO())
    console.print(tree)
    rich_str = console.file.getvalue()
    return rich_str


def to_multiindex_df(experiments) -> pd.DataFrame:
    """Excel-style merged headers"""
    test_names = sorted({r.name for exp in experiments for r in exp["tests"]})
    metrics = ["status"] + sorted(
        {k for exp in experiments for r in exp["tests"] for k in r.intermediates or {}}
    )

    columns = pd.MultiIndex.from_product([test_names, metrics], names=["test", "metric"])
    data = {}

    for exp in experiments:
        row_data = pd.Series(index=columns, dtype=object)
        for result in exp["tests"]:
            row_data[(result.name, "status")] = result.status
            for metric, value in result.intermediates.items():
                row_data[(result.name, metric)] = str(value)
        data[exp["name"]] = row_data

    return pd.DataFrame(data).T


def simple_table(experiments: List[Experiment]):
    rows = []
    all_tests = set(t.name for exp in experiments for t in exp.tests)

    for exp in experiments:
        test_dict = {t.name: t for t in exp.tests}
        row = [exp.name]
        for test_name in sorted(all_tests):
            test = test_dict.get(test_name)
            if test:
                row.extend([test.status, str(test.intermediates or {})])
            else:
                row.extend(["", ""])
        rows.append(row)

    headers = (
        ["exp"]
        + [f"{t}_status" for t in sorted(all_tests)]
        + [f"{t}_metrics" for t in sorted(all_tests)]
    )
    print(tabulate(rows, headers=headers, tablefmt="grid"))


def interactive_table(df: pd.DataFrame):
    fig = go.Figure(
        data=[
            go.Table(
                columnwidth=[200, 100, 200],
                header=dict(values=list(df.columns), height=40),
                cells=dict(values=[df[col] for col in df.columns]),
            )
        ]
    )
    fig.show()


class TableBuilder:
    def __init__(self):
        self.experiments: List[Experiment] = []

    def add_experiment(self, name: str, results: List[AnalysisResult]) -> Self:
        # Store as dict with name and tests
        self.experiments.append({"name": name, "tests": results})
        return self

    def tree(self) -> str:
        return print_tree(self.experiments)

    def _simple_df(self) -> pd.DataFrame:
        """Flat table: experiment | test_name | status | intermediates"""
        rows = []
        for exp in self.experiments:
            for result in exp["tests"]:
                row = {"experiment": exp["name"], "test": result.name, "status": result.status}
                if result.intermediates:
                    row.update(result.intermediates)
                rows.append(row)
        return pd.DataFrame(rows)

    def dataframe(self, multiindex: bool = True) -> pd.DataFrame:
        return (
            to_multiindex_df(self.experiments) if multiindex else self.simple_df(self.experiments)
        )

    def csv(self, path: str):
        self._simple_df().to_csv(path, index=False)

    def excel(self, path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # Build headers dynamically
        col = 2  # Start at column B (A = experiments)
        test_metrics = {}  # test_name -> list of its metrics

        # Collect all test+metrics first
        for exp in self.experiments:
            for result in exp["tests"]:
                test_name = result.name
                if test_name not in test_metrics:
                    test_metrics[test_name] = ["status"]
                if result.intermediates:
                    test_metrics[test_name].extend(result.intermediates.keys())

        # Write headers (starting column B)
        for test_name, metrics in test_metrics.items():
            # Row 1: Test name (merged)
            start_col = get_column_letter(col)
            end_col = get_column_letter(col + len(metrics) - 1)
            ws.merge_cells(f"{start_col}1:{end_col}1")
            ws[f"{start_col}1"] = test_name
            ws[f"{start_col}1"].font = Font(bold=True)
            ws[f"{start_col}1"].alignment = Alignment(horizontal="center")

            # Row 2: Metrics
            for i, metric in enumerate(metrics):
                ws.cell(row=2, column=col + i, value=metric)
                ws.cell(row=2, column=col + i).font = Font(bold=True)

            col += len(metrics)

        # Column A header
        ws["A1"] = "Experiment"
        ws["A1"].font = Font(bold=True)
        ws["A2"] = "Name"
        ws["A2"].font = Font(bold=True)

        # Write data rows
        for exp_idx, exp in enumerate(self.experiments, 3):
            # Column A: Experiment name
            ws.cell(row=exp_idx, column=1, value=exp["name"])
            ws.cell(row=exp_idx, column=1).font = Font(bold=True)

            col = 2  # Start at column B
            for result in exp["tests"]:
                test_metrics_list = test_metrics[result.name]

                # Status always first
                ws.cell(row=exp_idx, column=col, value=result.status)

                # Only this test's actual metrics
                col_idx = col + 1
                for metric in test_metrics_list[1:]:  # Skip status
                    value = result.intermediates.get(metric) if result.intermediates else None
                    ws.cell(row=exp_idx, column=col_idx, value=str(value) if value else "")
                    col_idx += 1

                col += len(test_metrics_list)

        # Auto-size columns and save
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(path)

    def excel_2(self, path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # Build headers dynamically
        col = 1
        test_metrics = {}  # test_name -> list of its metrics

        # Collect all test+metrics first
        for exp in self.experiments:
            for result in exp["tests"]:
                test_name = result.name
                if test_name not in test_metrics:
                    test_metrics[test_name] = ["status"]
                if result.intermediates:
                    test_metrics[test_name].extend(result.intermediates.keys())

        # Write headers
        for test_name, metrics in test_metrics.items():
            # Row 1: Test name (merged)
            start_col = get_column_letter(col)
            end_col = get_column_letter(col + len(metrics) - 1)
            ws.merge_cells(f"{start_col}1:{end_col}1")
            ws[f"{start_col}1"] = test_name
            ws[f"{start_col}1"].font = Font(bold=True)

            # Row 2: Metrics
            for i, metric in enumerate(metrics):
                ws.cell(row=2, column=col + i, value=metric)

            col += len(metrics)

        # Write data rows
        for exp_idx, exp in enumerate(self.experiments, 3):
            col = 1
            for result in exp["tests"]:
                test_metrics_list = test_metrics[result.name]

                # Status always first
                ws.cell(row=exp_idx, column=col, value=result.status)

                # Only this test's actual metrics
                col_idx = col + 1
                for metric in test_metrics_list[1:]:  # Skip status
                    value = result.intermediates.get(metric) if result.intermediates else None
                    ws.cell(row=exp_idx, column=col_idx, value=str(value) if value else "")
                    col_idx += 1

                col += len(test_metrics_list)

        # Auto-size columns and save
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(path)
